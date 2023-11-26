from cal_fp import calculate_forward_price

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

file = "input.xlsx"
his_price_df = pd.read_excel(file)

# To calculate the forward price
forward_prices = []
for  index, row in his_price_df.iterrows():
   cal = calculate_forward_price(row['Gold Spot Price'], row['Risk-free interest rate'], row['Time'])
   forward_prices.append(cal)

# Update datatable 
his_price_df = his_price_df.assign(Forward_price = forward_prices )
his_price_df['Date'] = pd.to_datetime(his_price_df['Date']).dt.year


# Calculate Profit and Loss
profit_n_loss = []
for index, row in his_price_df.iterrows():
    if index==0:
        cal = 0      
    else: 
        cal = his_price_df['Gold Spot Price'].iloc[index] - his_price_df['Forward_price'].iloc[index-1]
    profit_n_loss.append(round(cal, 2))

 # Update datatable 
his_price_df = his_price_df.assign(Profit_n_loss = profit_n_loss )  

his_price_df.rename(columns = {'Profit_n_loss':'Profit and loss'}, inplace = True)
his_price_df.rename(columns = {'Forward_price':'Forward price'}, inplace = True)
his_price_df.rename(columns = {'Date':'Year'}, inplace = True)

# Export data to excel
writer = pd.ExcelWriter("output.xlsx", engine='openpyxl')
his_price_df.to_excel(writer, sheet_name='Sheet1', index=False)
ws = writer.sheets['Sheet1']

# Format the data table
for column in ws.iter_cols(min_row=1, min_col=1, max_row=1+len(his_price_df.index), max_col=6):
    name = get_column_letter(column[0].column)
    # set witdth
    new_col_length = max(len(str(cell.value)) for cell in column)
    ws.column_dimensions[name].width = new_col_length+2 
    # set border to cell
    for cell in column:
        cell.border = Border(top=Side(style='thin'),
                            bottom=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin'))
# Draw chart
# BarChart - Forward prices
chart_fp = BarChart()
chart_fp.title = "The forward price and P&L of Gold during 2002-2022"
chart_fp.x_axis.title = 'Year'
chart_fp.y_axis.title = 'Forward price'
chart_fp.type = 'col'

y_fp = Reference(ws, min_col=5, min_row=1, max_row=1+len(his_price_df.index), max_col=5)
chart_fp.add_data(y_fp, titles_from_data=True)

# LineChart - Profit and Loss
chart_pnl = LineChart()
y_pnl = Reference(ws, min_col=6, min_row=1, max_row=1+len(his_price_df.index), max_col=6)
x_values = Reference(ws, min_col=1, min_row=2, max_row=1+len(his_price_df.index))
chart_pnl.add_data(y_pnl, titles_from_data=True)
chart_pnl.set_categories(x_values)
chart_pnl.y_axis.title = "PnL"


chart_fp += chart_pnl
ws.add_chart(chart_fp, "H4") 

writer.close()

